import os
import shutil
import calendar
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils import get_column_letter

month = [None, 'янв.', 'фев.', 'мар.', 'апр.', 'май.', 'июн.', 'июл.', 'авг.', 'сен. ', 'окт.', 'ноя.', 'дек.']
original_header = ['Порядковый №', 'Сотрудник (Посетитель)', 'Дата', 'Время', 'Подразделение', 'Событие',
                   'Устройство', 'Помещение', 'Пользователь', 'Категория события', 'Подкатегория события',
                   'Дата и время записи']
date = calendar.Calendar()
contractors = {}


def parsing(filename, day_month):
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    rows = sheet.max_row
    columns = sheet.max_column
    header = [row.value for row in sheet[4]]
    if header == original_header:
        for row in range(5, rows + 1):
            string = ''
            for column in range(2, columns + 1):
                cell = sheet.cell(row=row, column=column)
                string = string + str(cell.value) + ';'
            string = string.split(';')
            sorted(string, day_month)
    else:
        print('Неправильный формат отчета')


def sorted(string, day_month):
    person = string[0]
    date = string[1]
    day = int(date.split('.')[0])
    contract = string[3]
    zone = string[6]
    if contract not in contractors.keys():
        contractors[contract] = {
            zone: {
                'total_month': 1,
                'total_day': [0] * day_month,
                date: {
                    'day_people': 1,
                    'person': [person]
                }
            }
        }
        contractors[contract][zone]['total_day'][day - 1] += 1
    else:
        if zone not in contractors[contract].keys():
            contractors[contract][zone] = {
                'total_month': 1,
                'total_day': [0] * day_month,
                date: {
                    'day_people': 1,
                    'person': [person]
                }
            }
            contractors[contract][zone]['total_day'][day - 1] += 1
        elif date not in contractors[contract][zone].keys():
            contractors[contract][zone][date] = {'day_people': 1, 'person': [person]}
            contractors[contract][zone]['total_month'] += 1
            contractors[contract][zone]['total_day'][day - 1] += 1
        else:
            if person not in contractors[contract][zone][date]['person']:
                contractors[contract][zone][date]['person'].append(person)
                contractors[contract][zone][date]['day_people'] += 1
                contractors[contract][zone]['total_month'] += 1
                contractors[contract][zone]['total_day'][day - 1] += 1



def create_file(filename, filepath, date_list):
    filename = filename.split('/')[1]

    font = Font(name='Times New Roman',
                size=12,
                bold=False,
                italic=False,
                vertAlign=None,
                underline='none',
                strike=False,
                color='FF000000')

    font_top = Font(name='Times New Roman',
                    size=12,
                    bold=True,
                    italic=False,
                    vertAlign=None,
                    underline='none',
                    strike=False,
                    color='FF000000')

    border = Border(left=Side(border_style='thin',
                              color='FF000000'),
                    right=Side(border_style='thin',
                               color='FF000000'),
                    top=Side(border_style='thin',
                             color='FF000000'),
                    bottom=Side(border_style='thin',
                                color='FF000000'),
                    diagonal=Side(border_style='thin',
                                  color='FF000000'),
                    diagonal_direction=0,
                    outline=Side(border_style='thin',
                                 color='FF000000'),
                    vertical=Side(border_style='thin',
                                  color='FF000000'),
                    horizontal=Side(border_style='thin',
                                    color='FF000000')
                    )

    align_center = Alignment(horizontal='center',
                             vertical='bottom',
                             text_rotation=0,
                             wrap_text=False,
                             shrink_to_fit=False,
                             indent=0)

    align_top = Alignment(horizontal='center',
                          vertical='center',
                          text_rotation=0,
                          wrap_text=False,
                          shrink_to_fit=False,
                          indent=0)

    wb = Workbook()
    ws = wb.active
    position = {}  # Временное хранилище позиций в таблице в зависимости от территории
    wb.remove(wb.active)  # Удаление дефолтного листа
    for contractor in contractors.keys():
        for zone in contractors[contractor].keys():
            col = 5
            if zone not in wb.sheetnames:  # Если в книге нет листа территории
                position[zone] = {'start_row': 5, 'start_column': 2, 'index': 1, 'day': 5}  # Обозначение стартовых позиций
                ws = wb.create_sheet(zone)

                ws['E2'] = 'Отчет нахождения на объекте {}'.format(filename)
                ws['E2'].font = font_top
                # ---------------#
                ws['B4'] = '№'
                ws.column_dimensions['B'].width = 5
                # ---------------#
                ws['C4'] = 'Подрядчик'
                ws.column_dimensions['C'].width = 35
                # ---------------#
                ws['D4'] = 'Человек за месяц'
                ws.column_dimensions['D'].width = 19
                # ---------------#
                for i in date_list:
                    ws.cell(row=4, column=col).value = i
                    col += 1
            ws = wb[zone]  # Выбрать активным лист согласно территории
            ws.cell(row=position[zone]['start_row'],
                    column=position[zone]['start_column']).value = position[zone]['index']
            ws.cell(row=position[zone]['start_row'],
                    column=position[zone]['start_column']+1).value = contractor
            ws.cell(row=position[zone]['start_row'],
                    column=position[zone]['start_column']+2).value = contractors[contractor][zone]['total_month']
            for day_value in contractors[contractor][zone]['total_day']:
                ws.cell(row=position[zone]['start_row'], column=position[zone]['day']).value = day_value
                position[zone]['day'] += 1
            position[zone]['day'] = 5
            position[zone]['index'] = position[zone]['index'] + 1
            position[zone]['start_row'] = position[zone]['start_row'] + 1

    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for row in range(4, ws.max_row+1):
            for column in range(2, ws.max_column+1):
                cell = ws.cell(row=row, column=column)
                cell.border = border
                cell.font = font
                if row >= 4 and column == 2:  # Форматирования столбца №
                    cell.alignment = align_center
                if row == 4 and column >= 2:
                    cell.font = font_top
                    cell.alignment = align_top
                if column >= 5:
                    i = get_column_letter(column)
                    ws.column_dimensions[i].width = 10
    wb.save(filepath)

def start():
    folders = [folder for folder in os.walk('Отчеты')]
    for folder in folders[1:]:
        for file in folder[2]:
            if file.endswith('.xlsx') and file.startswith('События') and not folder[0].endswith('old'):
                file_month = month.index(file.split(' ')[-3])
                file_year = int(file.split(' ')[-2])
                date_list = []
                for i in date.itermonthdates(file_year, file_month):
                    if i.month == file_month:
                        date_list.append(i.strftime('%d.%m.%Y'))
                perco_file = os.path.join('', folder[0], file)
                print('Файл {}'.format(perco_file))
                parsing(perco_file, len(date_list))
                create_file(folder[0],
                            os.path.join('', folder[0], 'Отчет нахождения на объекте {}'.format(' '.join(file.split(' ')[1:]))),
                            date_list)
                contractors.clear()
                shutil.copy2(perco_file, os.path.join(folder[0], 'old', file))
                os.remove(perco_file)


if __name__ == "__main__":
    start()

